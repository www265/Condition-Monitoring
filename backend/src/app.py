from flask import Flask, request, jsonify, send_from_directory, render_template, url_for
from werkzeug.utils import secure_filename
from flask_cors import CORS
import os
import pandas as pd
import logging
import numpy as np
from datetime import datetime
import openpyxl  # 用于读取 .xlsx 文件
import xlrd  # 用于读取 .xls 文件
from signals import generate_sine_wave, generate_square_wave, generate_triangle_wave
from wavefftgse import SignalProcessor  # 导入 SignalProcessor 类
from DimentionReduction import DimentionReduction
import matplotlib.pyplot as plt
import io
import base64
import json

app = Flask(__name__)
CORS(app, resources={
    r"/api/*": {"origins": "http://localhost:8080"},
    r"/uploads/*": {"origins": "http://localhost:8080"}
})

# 配置上传目录和允许的文件扩展名
UPLOAD_FOLDER = './uploads'
ALLOWED_EXTENSIONS = {'csv', 'txt', 'xls', 'xlsx', 'dat'}
MAX_CONTENT_LENGTH = 10 * 1024 * 1024  # 限制文件大小为10MB
dim_red = DimentionReduction()

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# 设置日志配置
logging.basicConfig(filename='app.log', level=logging.INFO)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        logging.error('No file part')
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']

    if file.filename == '':
        logging.error('No selected file')
        return jsonify({'error': 'No selected file'}), 400

    if file and allowed_file(file.filename):
        try:
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)

            # 根据文件扩展名选择解析方法
            file_extension = filename.lower().rsplit('.', 1)[1]
            data_info = None

            if file_extension in ('csv', 'txt'):
                try:
                    df = pd.read_csv(file_path)
                    data_info = {'shape': df.shape}
                except Exception as e:
                    logging.error(f"Error parsing CSV/TXT file {filename}: {str(e)}")
                    return jsonify({'error': f'Failed to parse CSV/TXT file: {str(e)}'}), 500
            elif file_extension == 'xlsx':
                try:
                    wb = openpyxl.load_workbook(file_path)
                    sheet_names = wb.sheetnames
                    data_info = {'sheet_names': sheet_names}
                except Exception as e:
                    logging.error(f"Error parsing XLSX file {filename}: {str(e)}")
                    return jsonify({'error': f'Failed to parse XLSX file: {str(e)}'}), 500
            elif file_extension == 'xls':
                try:
                    wb = xlrd.open_workbook(file_path)
                    sheet_names = wb.sheet_names()
                    data_info = {'sheet_names': sheet_names}
                except Exception as e:
                    logging.error(f"Error parsing XLS file {filename}: {str(e)}")
                    return jsonify({'error': f'Failed to parse XLS file: {str(e)}'}), 500
            elif file_extension == 'dat':
                # 对于 .dat 文件，假设它是二进制数据，这里需要根据实际情况解析
                try:
                    with open(file_path, 'rb') as f:
                        bytes_data = f.read()
                    data = np.frombuffer(bytes_data, dtype=np.float64)  # 根据实际情况调整dtype
                except Exception as e:
                    logging.error(f"Error parsing DAT file {filename}: {str(e)}")
                    return jsonify({'error': f'Failed to parse DAT file: {str(e)}'}), 500
            else:
                raise ValueError("Unsupported file extension")

            logging.info(f"File {filename} uploaded successfully at {datetime.now()}")
            response_data = {
                'message': 'File uploaded successfully',
                'filename': filename,
                'file_path': f"/uploads/{filename}",  # 或者你需要的实际路径
                **(data_info or {})
            }
            return jsonify(response_data), 200
        except Exception as e:
            logging.error(f"Error processing file: {str(e)}")
            return jsonify({'error': str(e)}), 500
    else:
        logging.error('File type not allowed')
        return jsonify({'error': 'File type not allowed'}), 400


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


@app.route('/api/generate-signal', methods=['POST'])
def generate_signal():
    try:
        data = request.get_json(silent=True)  # silent=True 避免抛出异常

        if not data:
            raise ValueError("Request body is empty or not valid JSON.")

        # 定义一个字典来映射信号类型到生成函数
        signal_generators = {
            'sine': generate_sine_wave,
            'square': generate_square_wave,
            'triangle': generate_triangle_wave,
        }

        # 检查所有必须的参数是否存在
        required_params = ['type', 'frequency', 'amplitude', 'phase', 'duration', 'sampleRate']
        signal_type = data['type'].lower()
        if signal_type == 'square':
            required_params.append('duty_cycle')

        missing_params = [param for param in required_params if param not in data]
        if missing_params:
            raise ValueError(f"Missing parameters: {', '.join(missing_params)}")

        # 从请求体中获取参数并且转换类型
        frequency = float(data['frequency'])
        amplitude = float(data['amplitude'])
        phase = float(data['phase'])
        duration = float(data['duration'])
        sample_rate = int(data['sampleRate'])

        duty_cycle = float(data.get('duty_cycle')) if signal_type == 'square' else None

        # 参数验证
        if frequency <= 0 or amplitude <= 0 or duration <= 0 or sample_rate <= 0:
            raise ValueError("Frequency, amplitude, duration, and sample rate must be positive numbers.")
        if signal_type == 'square' and (duty_cycle is None or duty_cycle < 0 or duty_cycle > 1):
            raise ValueError("Duty cycle must be between 0 and 1 for square waves.")

        # 根据信号类型生成信号
        generator_func = signal_generators.get(signal_type)
        if not generator_func:
            raise ValueError("Unsupported signal type")

        if signal_type == 'square':
            signal = generator_func(frequency, amplitude, duty_cycle, duration, sample_rate)
        else:
            signal = generator_func(frequency, amplitude, phase, duration, sample_rate)

        logging.info(
            f"Generated {signal_type} wave with parameters: freq={frequency}, amp={amplitude}, phase={phase}, duration={duration}, sampleRate={sample_rate}"
        )

        return jsonify({
            'message': f'{signal_type.capitalize()} signal generated successfully',
            'signal': signal.tolist(),
            'type': signal_type
        }), 200

    except ValueError as ve:
        logging.error(f"Client error generating signal: {str(ve)}")
        return jsonify({'error': str(ve)}), 400
    except Exception as e:
        logging.error(f"Server error generating signal: {str(e)}", exc_info=True)  # 记录完整的堆栈跟踪
        return jsonify({'error': "An unexpected error occurred."}), 500

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def catch_all(path):
    if path != "" and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    else:
        return send_from_directory(app.static_folder, 'index.html')


@app.route('/api/process-signal', methods=['POST'])
def process_signal():
    try:
        if request.is_json:
            json_data = request.get_json()
            logging.info(f"Received JSON data: {json.dumps(json_data, indent=2)}")

            if not json_data or 'data' not in json_data or 'sample_rate' not in json_data or 'high_pass_cutoff' not in json_data:
                return jsonify({'error': 'Invalid input'}), 400

            data = json_data['data']
            sample_rate = float(json_data['sample_rate'])
            high_pass_cutoff = float(json_data['high_pass_cutoff'])

            # 将 data 转换为 NumPy 数组
            data['x'] = np.array(data.get('x', []))
            data['y'] = np.array(data.get('y', []))

            processor = SignalProcessor(data=data, sample_rate=sample_rate, high_pass_cutoff=high_pass_cutoff)
            results = processor.process_all(high_pass_cutoff=high_pass_cutoff)
            logging.info(f"Processed results: {json.dumps(results, indent=2)}")

            response_data = {
                'message': 'Signal processed successfully',
                'results': results,
            }

            return jsonify(response_data), 200

        elif 'file' in request.files:
            file = request.files['file']
            sample_rate = float(request.form.get('sampleRate', 1000))
            high_pass_cutoff = float(request.form.get('highPassCutoff', 100))

            if file.filename == '':
                logging.error('No selected file')
                return jsonify({'error': 'No selected file'}), 400

            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)

                # 加载数据到 NumPy 数组中
                data = load_data(file_path)

                processor = SignalProcessor(data=data, sample_rate=sample_rate, high_pass_cutoff=high_pass_cutoff)
                results = processor.process_all(high_pass_cutoff=high_pass_cutoff)

                response_data = {
                    'message': 'Signal processed successfully',
                    'results': results,
                }

                return jsonify(response_data), 200

            else:
                logging.error('File type not allowed')
                return jsonify({'error': 'File type not allowed'}), 400

        else:
            return jsonify({'error': 'Unsupported request format'}), 400

    except Exception as e:
        logging.error(f"Error processing signal: {str(e)}")
        return jsonify({'error': str(e)}), 500


def load_data(file_path):
    """根据文件扩展名加载数据"""
    file_extension = file_path.lower().rsplit('.', 1)[1]

    if file_extension in ('csv', 'txt'):
        try:
            df = pd.read_csv(file_path)
            # 确保 X 和 Y 列只包含数值数据
            df['X'] = pd.to_numeric(df['X'], errors='coerce')
            df['Y'] = pd.to_numeric(df['Y'], errors='coerce')
            df = df.dropna(subset=['X', 'Y'])

            if df.empty:
                raise ValueError("No valid numeric data found in the file")

            data = {'x': df['X'].values, 'y': df['Y'].values}
        except Exception as e:
            logging.error(f"Error parsing CSV/TXT file {file_path}: {str(e)}")
            raise ValueError(f"Failed to parse CSV/TXT file: {str(e)}")
    elif file_extension == 'xlsx':
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            data = {'x': [], 'y': []}
            for row in sheet.iter_rows(min_row=2, values_only=True):  # 假设数据从第二行开始
                x, y = row[:2]  # 假设 X 和 Y 数据位于前两列
                if isinstance(x, (int, float)) and isinstance(y, (int, float)):
                    data['x'].append(x)
                    data['y'].append(y)
            data['x'] = np.array(data['x'])
            data['y'] = np.array(data['y'])
        except Exception as e:
            logging.error(f"Error parsing XLSX file {file_path}: {str(e)}")
            raise ValueError(f"Failed to parse XLSX file: {str(e)}")
    elif file_extension == 'xls':
        try:
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_index(0)
            data = {'x': [], 'y': []}
            for i in range(1, sheet.nrows):  # 假设数据从第二行开始
                x, y = sheet.row_values(i)[:2]  # 假设 X 和 Y 数据位于前两列
                if isinstance(x, (int, float)) and isinstance(y, (int, float)):
                    data['x'].append(x)
                    data['y'].append(y)
            data['x'] = np.array(data['x'])
            data['y'] = np.array(data['y'])
        except Exception as e:
            logging.error(f"Error parsing XLS file {file_path}: {str(e)}")
            raise ValueError(f"Failed to parse XLS file: {str(e)}")
    elif file_extension == 'dat':
        # 对于 .dat 文件，假设它是二进制数据，这里需要根据实际情况解析
        try:
            with open(file_path, 'rb') as f:
                bytes_data = f.read()
            data = np.frombuffer(bytes_data, dtype=np.float64)  # 根据实际情况调整dtype
        except Exception as e:
            logging.error(f"Error parsing DAT file {file_path}: {str(e)}")
            raise ValueError(f"Failed to parse DAT file: {str(e)}")
    else:
        raise ValueError("Unsupported file extension")

    return data


def generate_plot(results):
    """根据结果生成图表并返回图表URL"""
    try:
        plt.plot(results['x'], results['y'])

        # 将图表保存为 PNG 图像，并将其编码为 Base64 字符串
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        img_str = base64.b64encode(buf.getvalue()).decode('utf-8').replace('\n', '')
        plt.close()

        plot_url = f"data:image/png;base64,{img_str}"
        return plot_url
    except Exception as e:
        logging.error(f"Error generating plot: {str(e)}")
        return None


@app.route('/plot')
def plot():
    plot_url = request.args.get('plot_url')
    if not plot_url:
        return "Plot not available", 400
    return render_template('plot.html', plot_url=plot_url)

@app.route('/api/dimention-reduction', methods=['POST'])
def dimention_reduction():
    try:
        data = request.get_json()

        if not data:
            raise ValueError("Request body is empty or not valid JSON.")

        algorithm = data.get('algorithm')
        dataset = data.get('dataset')
        parameters = data.get('parameters', {})

        result = dim_red.apply_dimentionality_reduction(algorithm, dataset, parameters)

        return jsonify(result), 200

    except ValueError as ve:
        logging.error(f"ValueError: {str(ve)}")
        return jsonify({'error': str(ve)}), 400
    except Exception as e:
        logging.error(f"Exception: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/show-data', methods=['POST'])
def show_data():
    try:
        data = request.get_json()

        if not data:
            raise ValueError("Request body is empty or not valid JSON.")

        dataset = data.get('dataset')

        X, y, feature_names, target_names = dim_red.load_dataset(dataset)

        # Prepare the data to be sent back to the client
        dataset_info = {
            'feature_names': feature_names,
            'target_names': target_names.tolist(),  # Convert NumPy array to list
            'data': X[:10].tolist(),  # Send only the first 10 samples for brevity
            'targets': y[:10].tolist()  # Convert NumPy array to list
        }

        return jsonify(dataset_info), 200

    except ValueError as ve:
        logging.error(f"ValueError: {str(ve)}")
        return jsonify({'error': str(ve)}), 400
    except Exception as e:
        logging.error(f"Exception: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True)



